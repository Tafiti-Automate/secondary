import csv
import hashlib
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config.crud import AuditedFormMixin, SoftDeleteView
from config.audit import record_audit
from config.mixins import ACADEMIC_MANAGERS, AcademicManagerRequiredMixin, AcademicStaffRequiredMixin
from config.tables import ModelTableView
from students.models import Student
from academics.models import ClassLevel, Stream, SubjectAllocation, Term
from .forms import ExamSessionForm, ExaminationForm, GradeBoundaryForm, GradeScaleForm, UNEBCandidateForm, UNEBCandidateSubjectForm, UNEBCAImportForm, UNEBContinuousAssessmentForm, UpperAssessmentComponentForm, UpperAssessmentPlanForm
from .models import ExamSession, Examination, ExaminationResult, GradeBoundary, GradeScale, UNEBCandidate, UNEBCandidateSubject, UNEBContinuousAssessment, UNEBExportBatch, UNEBIntegrationAdapter, UpperAssessmentComponent, UpperAssessmentPlan
from .readiness import active_uneb_adapter, uneb_export_columns, uneb_export_preflight, uneb_export_row
from .services import process_upper_subject_results


def teacher_examinations(user):
    qs = Examination.objects.filter(is_deleted=False).select_related("session", "term", "subject", "stream", "created_by")
    if user.is_superuser or user.role in ACADEMIC_MANAGERS:
        return qs
    if user.role == "teacher":
        allocation = SubjectAllocation.objects.filter(teacher=user, subject_id=OuterRef("subject_id"), stream_id=OuterRef("stream_id"), term_id=OuterRef("term_id"), is_deleted=False, is_active=True)
        return qs.annotate(is_allocated=Exists(allocation)).filter(is_allocated=True)
    return qs.none()


class ExaminationListView(AcademicStaffRequiredMixin, ListView):
    template_name = "exams/exam_list.html"
    context_object_name = "examinations"
    paginate_by = 20

    def get_queryset(self):
        qs = teacher_examinations(self.request.user)
        if self.request.GET.get("status"):
            qs = qs.filter(status=self.request.GET["status"])
        return qs


class ExaminationCreateView(AcademicStaffRequiredMixin, AuditedFormMixin, CreateView):
    model = Examination
    form_class = ExaminationForm
    template_name = "shared/form.html"
    success_url = reverse_lazy("exams:list")
    audit_action = "create"
    extra_context = {"page_title": "Create examination", "eyebrow": "Examinations", "submit_label": "Create examination"}

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == "teacher" and not self.request.user.is_superuser:
            allocations = SubjectAllocation.objects.filter(teacher=self.request.user, is_deleted=False, is_active=True)
            form.fields["subject"].queryset = form.fields["subject"].queryset.filter(allocations__in=allocations).distinct()
            form.fields["stream"].queryset = form.fields["stream"].queryset.filter(subject_allocations__in=allocations).distinct()
            form.fields["term"].queryset = form.fields["term"].queryset.filter(subject_allocations__in=allocations).distinct()
            form.fields["status"].choices = [("draft", "Draft"), ("entry", "Mark entry")]
        form.fields["component"].queryset = UpperAssessmentComponent.objects.filter(
            plan__status__in=("approved", "locked"), is_active=True, is_deleted=False,
        ).select_related("plan__subject", "plan__class_level")
        return form


class ExaminationUpdateView(AcademicStaffRequiredMixin, AuditedFormMixin, UpdateView):
    model = Examination
    form_class = ExaminationForm
    template_name = "shared/form.html"
    success_url = reverse_lazy("exams:list")
    audit_action = "update"
    extra_context = {"page_title": "Edit examination", "eyebrow": "Examinations", "submit_label": "Save changes"}

    def get_queryset(self):
        return teacher_examinations(self.request.user).exclude(status="locked")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if self.request.user.role == "teacher" and not self.request.user.is_superuser:
            form.fields["status"].choices = [("draft", "Draft"), ("entry", "Mark entry")]
        form.fields["component"].queryset = UpperAssessmentComponent.objects.filter(
            plan__status__in=("approved", "locked"), is_active=True, is_deleted=False,
        ).select_related("plan__subject", "plan__class_level")
        return form


class UpperAssessmentPlanListView(AcademicManagerRequiredMixin, ListView):
    template_name = "exams/upper_plan_list.html"
    context_object_name = "plans"
    paginate_by = 30

    def get_queryset(self):
        return UpperAssessmentPlan.objects.filter(is_deleted=False).select_related(
            "academic_year", "class_level", "subject", "grade_scale", "approved_by"
        ).prefetch_related("components")


class UpperAssessmentPlanCreateView(AcademicManagerRequiredMixin, AuditedFormMixin, CreateView):
    model = UpperAssessmentPlan
    form_class = UpperAssessmentPlanForm
    template_name = "shared/form.html"
    audit_action = "create"
    extra_context = {"page_title": "Create S5/S6 assessment plan", "eyebrow": "Aligned Advanced Secondary", "submit_label": "Create draft plan"}

    def form_valid(self, form):
        form.instance.status = "draft"
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("exams:upper-plan-detail", args=[self.object.pk])


class UpperAssessmentPlanDetailView(AcademicManagerRequiredMixin, DetailView):
    template_name = "exams/upper_plan_detail.html"
    context_object_name = "plan"

    def get_queryset(self):
        return UpperAssessmentPlan.objects.filter(is_deleted=False).select_related(
            "academic_year", "class_level", "subject", "grade_scale", "approved_by"
        ).prefetch_related("components", "processed_results__student", "processed_results__term")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["terms"] = Term.objects.filter(
            academic_year=self.object.academic_year, number=3, is_deleted=False,
        )
        context["streams"] = Stream.objects.filter(
            academic_year=self.object.academic_year,
            class_level=self.object.class_level,
            is_deleted=False,
        )
        return context


class UpperAssessmentComponentCreateView(AcademicManagerRequiredMixin, AuditedFormMixin, CreateView):
    model = UpperAssessmentComponent
    form_class = UpperAssessmentComponentForm
    template_name = "shared/form.html"
    audit_action = "create"
    extra_context = {"page_title": "Add S5/S6 assessment component", "eyebrow": "Aligned annual plan", "submit_label": "Add component"}

    def dispatch(self, request, *args, **kwargs):
        self.plan = get_object_or_404(UpperAssessmentPlan, pk=kwargs["pk"], status="draft", is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.plan = self.plan
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("exams:upper-plan-detail", args=[self.plan.pk])


class UpperAssessmentComponentUpdateView(AcademicManagerRequiredMixin, AuditedFormMixin, UpdateView):
    model = UpperAssessmentComponent
    form_class = UpperAssessmentComponentForm
    template_name = "shared/form.html"
    audit_action = "update"
    extra_context = {"page_title": "Edit S5/S6 assessment component", "eyebrow": "Aligned annual plan", "submit_label": "Save component"}

    def get_queryset(self):
        return UpperAssessmentComponent.objects.filter(plan__status="draft", is_deleted=False)

    def get_success_url(self):
        return reverse("exams:upper-plan-detail", args=[self.object.plan_id])


class UpperAssessmentPlanWorkflowView(AcademicManagerRequiredMixin, View):
    def post(self, request, pk, operation):
        plan = get_object_or_404(UpperAssessmentPlan, pk=pk, is_deleted=False)
        try:
            if operation == "approve" and plan.status == "draft":
                plan.status, plan.approved_by = "approved", request.user
            elif operation == "lock" and plan.status == "approved":
                plan.status, plan.approved_by = "locked", plan.approved_by or request.user
            elif operation == "reopen" and plan.status in {"approved", "locked"}:
                plan.status = "draft"
                plan.approved_by = None
                plan.approved_at = None
            else:
                raise ValidationError("That assessment-plan transition is not allowed.")
            plan.save()
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            record_audit(request, f"upper_plan_{operation}", plan, f"S5/S6 assessment plan {operation}")
            messages.success(request, f"Assessment plan {operation} completed.")
        return redirect("exams:upper-plan-detail", pk=pk)


class UpperAssessmentPlanProcessView(AcademicManagerRequiredMixin, View):
    def post(self, request, pk):
        plan = get_object_or_404(UpperAssessmentPlan, pk=pk, is_deleted=False)
        try:
            term = Term.objects.get(pk=request.POST.get("term"), is_deleted=False)
            stream = Stream.objects.get(pk=request.POST.get("stream"), is_deleted=False)
            results = process_upper_subject_results(plan, term, stream, request.user)
        except (Term.DoesNotExist, Stream.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Choose a valid Term III and stream.")
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
        else:
            record_audit(request, "upper_plan_process", plan, f"Processed {len(results)} S5/S6 subject result(s)")
            messages.success(request, f"Processed {len(results)} S5/S6 subject result(s).")
        return redirect("exams:upper-plan-detail", pk=pk)


class ExaminationDetailView(AcademicStaffRequiredMixin, DetailView):
    template_name = "exams/exam_detail.html"
    context_object_name = "examination"

    def get_queryset(self):
        return teacher_examinations(self.request.user).prefetch_related("results__student")


class ExaminationDeleteView(AcademicManagerRequiredMixin, SoftDeleteView):
    model = Examination
    success_url_name = "exams:list"


class ExamMarkEntryView(AcademicStaffRequiredMixin, View):
    template_name = "exams/mark_entry.html"

    def examination(self, request, pk):
        return get_object_or_404(teacher_examinations(request.user), pk=pk)

    def get(self, request, pk):
        exam = self.examination(request, pk)
        students = Student.objects.filter(stream=exam.stream, is_deleted=False, is_active=True)
        existing = {x.student_id: x for x in ExaminationResult.objects.filter(examination=exam, is_deleted=False)}
        return render(request, self.template_name, {"examination": exam, "rows": [{"student": s, "result": existing.get(s.pk)} for s in students], "moderation": False})

    @transaction.atomic
    def post(self, request, pk):
        exam = self.examination(request, pk)
        if exam.locked or (exam.status not in {"draft", "entry"} and request.user.role not in ACADEMIC_MANAGERS):
            messages.error(request, "Mark entry is closed for this examination.")
            return redirect("exams:detail", pk=pk)
        saved, errors = 0, []
        for student in Student.objects.filter(stream=exam.stream, is_deleted=False, is_active=True):
            raw = request.POST.get(f"score_{student.pk}", "").strip()
            if not raw:
                continue
            try:
                result, _ = ExaminationResult.objects.get_or_create(examination=exam, student=student, defaults={"score": Decimal(raw), "marked_by": request.user})
                result.score = Decimal(raw)
                result.teacher_remark = request.POST.get(f"remark_{student.pk}", "").strip()
                result.marked_by = request.user
                result.save()
                saved += 1
            except (InvalidOperation, ValidationError) as exc:
                errors.append(f"{student.full_name}: {exc}")
        if errors:
            messages.error(request, "Some marks were rejected: " + "; ".join(errors[:3]))
        messages.success(request, f"Saved {saved} examination mark(s).")
        if saved:
            record_audit(request, "mark_entry", exam, f"Entered {saved} examination mark(s)", {"saved": saved})
        return redirect("exams:marks", pk=pk)


class ExamModerationView(AcademicManagerRequiredMixin, View):
    template_name = "exams/mark_entry.html"

    def get(self, request, pk):
        exam = get_object_or_404(Examination, pk=pk, is_deleted=False)
        rows = [{"student": result.student, "result": result} for result in exam.results.filter(is_deleted=False).select_related("student")]
        return render(request, self.template_name, {"examination": exam, "rows": rows, "moderation": True})

    @transaction.atomic
    def post(self, request, pk):
        exam = get_object_or_404(Examination, pk=pk, is_deleted=False)
        if exam.locked:
            messages.error(request, "A locked examination cannot be moderated.")
            return redirect("exams:detail", pk=pk)
        for result in exam.results.filter(is_deleted=False):
            raw = request.POST.get(f"score_{result.student_id}", "").strip()
            if raw:
                result.moderated_score = Decimal(raw)
                result.moderated_by = request.user
                result.save()
        exam.transition("moderate", request.user)
        record_audit(request, "moderate", exam, "Moderated examination results")
        messages.success(request, "Moderated scores saved and examination moved to moderation complete.")
        return redirect("exams:detail", pk=pk)


class ExaminationTransitionView(AcademicStaffRequiredMixin, View):
    allowed_states = {
        "submit": {"draft", "entry"}, "moderate": {"submitted"}, "approve": {"moderated"},
        "publish": {"approved"}, "lock": {"published"}, "reopen": {"submitted", "moderated", "approved", "published"},
    }

    def post(self, request, pk, operation):
        exam = get_object_or_404(teacher_examinations(request.user), pk=pk)
        if operation not in self.allowed_states or exam.status not in self.allowed_states[operation]:
            messages.error(request, "That workflow transition is not valid from the current status.")
        elif operation != "submit" and request.user.role not in ACADEMIC_MANAGERS and not request.user.is_superuser:
            messages.error(request, "Only academic management can perform that approval action.")
        else:
            exam.transition(operation, request.user)
            if operation in {"approve", "publish", "lock"}:
                exam.results.filter(is_deleted=False).update(approved=True)
            messages.success(request, f"Examination {operation} action completed.")
            record_audit(request, operation, exam, f"Examination workflow action: {operation}")
        return redirect("exams:detail", pk=pk)


class UNEBHubView(AcademicStaffRequiredMixin, View):
    def get(self, request):
        candidates = UNEBCandidate.objects.filter(is_deleted=False)
        records = UNEBContinuousAssessment.objects.filter(is_deleted=False)
        if request.user.role == "teacher":
            records = records.filter(candidate_subject__subject__allocations__teacher=request.user, candidate_subject__subject__allocations__is_deleted=False).distinct()
        return render(request, "exams/uneb_hub.html", {
            "candidate_count": candidates.count() if request.user.role in ACADEMIC_MANAGERS or request.user.is_superuser else None,
            "record_count": records.count(), "draft_count": records.filter(status="draft").count(),
            "verified_count": records.filter(status="verified").count(), "approved_count": records.filter(status="approved").count(),
            "recent_batches": UNEBExportBatch.objects.filter(is_deleted=False)[:5] if request.user.role in ACADEMIC_MANAGERS or request.user.is_superuser else [],
        })


class UNEBCandidateListView(AcademicManagerRequiredMixin, ListView):
    template_name = "exams/uneb_candidate_list.html"
    context_object_name = "candidates"
    paginate_by = 25

    def get_queryset(self):
        qs = UNEBCandidate.objects.filter(is_deleted=False).select_related("student").annotate(subject_count=Count("registered_subjects", distinct=True), ca_count=Count("registered_subjects__continuous_assessments", distinct=True)).order_by("examination_year", "candidate_name")
        if self.request.GET.get("year"):
            qs = qs.filter(examination_year=self.request.GET["year"])
        if self.request.GET.get("level"):
            qs = qs.filter(examination_level=self.request.GET["level"])
        if self.request.GET.get("q"):
            qs = qs.filter(Q(candidate_name__icontains=self.request.GET["q"]) | Q(index_number__icontains=self.request.GET["q"]) | Q(student__admission_number__icontains=self.request.GET["q"]))
        return qs


class UNEBCandidateCreateView(AcademicManagerRequiredMixin, AuditedFormMixin, CreateView):
    model = UNEBCandidate
    form_class = UNEBCandidateForm
    template_name = "shared/form.html"
    audit_action = "create"
    extra_context = {"page_title": "Create UNEB candidate record", "eyebrow": "Pre-submission candidate register", "submit_label": "Save candidate"}

    def get_success_url(self):
        return reverse_lazy("exams:uneb-candidate-detail", kwargs={"pk": self.object.pk})


class UNEBCandidateDetailView(AcademicManagerRequiredMixin, DetailView):
    model = UNEBCandidate
    queryset = UNEBCandidate.objects.filter(is_deleted=False).select_related("student").prefetch_related("registered_subjects__subject", "registered_subjects__continuous_assessments__class_level", "registered_subjects__continuous_assessments__term")
    template_name = "exams/uneb_candidate_detail.html"
    context_object_name = "candidate"


class UNEBCandidateSubjectCreateView(AcademicManagerRequiredMixin, AuditedFormMixin, CreateView):
    model = UNEBCandidateSubject
    form_class = UNEBCandidateSubjectForm
    template_name = "shared/form.html"
    audit_action = "create"
    extra_context = {"page_title": "Register UNEB subject", "eyebrow": "Candidate subject register", "submit_label": "Register subject"}

    def get_initial(self):
        initial = super().get_initial()
        initial["candidate"] = self.request.GET.get("candidate")
        return initial

    def get_success_url(self):
        return reverse_lazy("exams:uneb-candidate-detail", kwargs={"pk": self.object.candidate_id})


def scoped_uneb_ca(user):
    qs = UNEBContinuousAssessment.objects.filter(is_deleted=False).select_related("candidate_subject__candidate__student", "candidate_subject__subject", "class_level", "term", "entered_by")
    if user.role == "teacher" and not user.is_superuser:
        return qs.filter(candidate_subject__subject__allocations__teacher=user, candidate_subject__subject__allocations__is_deleted=False).distinct()
    return qs


class UNEBCAListView(AcademicStaffRequiredMixin, ListView):
    template_name = "exams/uneb_ca_list.html"
    context_object_name = "records"
    paginate_by = 30

    def get_queryset(self):
        qs = scoped_uneb_ca(self.request.user)
        if self.request.GET.get("status"):
            qs = qs.filter(status=self.request.GET["status"])
        if self.request.GET.get("year"):
            qs = qs.filter(candidate_subject__candidate__examination_year=self.request.GET["year"])
        return qs


class UNEBCACreateView(AcademicStaffRequiredMixin, AuditedFormMixin, CreateView):
    model = UNEBContinuousAssessment
    form_class = UNEBContinuousAssessmentForm
    template_name = "shared/form.html"
    success_url = reverse_lazy("exams:uneb-ca-list")
    audit_action = "create"
    extra_context = {"page_title": "Enter UNEB continuous assessment", "eyebrow": "S3/S4 or S5/S6 evidence register", "submit_label": "Save draft score"}

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["status"].choices = [("draft", "Draft")]
        if self.request.user.role == "teacher" and not self.request.user.is_superuser:
            form.fields["candidate_subject"].queryset = form.fields["candidate_subject"].queryset.filter(subject__allocations__teacher=self.request.user, subject__allocations__is_deleted=False).distinct()
            form.fields["status"].choices = [("draft", "Draft")]
        return form

    def get_initial(self):
        initial = super().get_initial()
        initial["candidate_subject"] = self.request.GET.get("candidate_subject")
        return initial

    def form_valid(self, form):
        form.instance.entered_by = self.request.user
        form.instance.status = "draft"
        return super().form_valid(form)


class UNEBCAWorkflowView(AcademicManagerRequiredMixin, View):
    @transaction.atomic
    def post(self, request, operation):
        ids = request.POST.getlist("record_ids")
        records = UNEBContinuousAssessment.objects.filter(pk__in=ids, is_deleted=False)
        now = timezone.now()
        if operation == "verify":
            records = records.filter(status="draft").exclude(evidence_reference="").exclude(candidate_subject__candidate__index_number="").exclude(candidate_subject__candidate__centre_number="")
            count = records.update(status="verified", verified_by=request.user, verified_at=now)
        elif operation == "approve":
            records = records.filter(status="verified")
            count = records.update(status="approved", approved_by=request.user, approved_at=now)
        else:
            count = 0
        messages.success(request, f"{count} UNEB continuous-assessment record(s) {operation}d.")
        return redirect("exams:uneb-ca-list")


class UNEBCAImportView(AcademicStaffRequiredMixin, View):
    template_name = "exams/uneb_ca_import.html"
    required = {"index_number", "subject_code", "class_level", "term_number", "component", "raw_score", "maximum_score", "evidence_reference"}

    def get(self, request):
        return render(request, self.template_name, {"form": UNEBCAImportForm()})

    def post(self, request):
        form = UNEBCAImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})
        try:
            reader = csv.DictReader(StringIO(form.cleaned_data["csv_file"].read().decode("utf-8-sig")))
            if not reader.fieldnames or not self.required.issubset(reader.fieldnames):
                raise ValueError("Required columns are missing. Use the current import template.")
            count = 0
            with transaction.atomic():
                for row_number, row in enumerate(reader, 2):
                    candidate = UNEBCandidate.objects.filter(index_number=row["index_number"].strip(), examination_level=form.cleaned_data["examination_level"], examination_year=form.cleaned_data["examination_year"], is_deleted=False).first()
                    if not candidate:
                        raise ValueError(f"Row {row_number}: candidate index number is not registered.")
                    registration = UNEBCandidateSubject.objects.filter(candidate=candidate, subject_code__iexact=row["subject_code"].strip(), is_deleted=False, is_registered=True).select_related("subject").first()
                    if not registration:
                        raise ValueError(f"Row {row_number}: subject is not registered for this candidate.")
                    class_level = ClassLevel.objects.filter(name__iexact=row["class_level"].strip(), is_deleted=False).first()
                    term = Term.objects.filter(academic_year__start_date__year__lte=form.cleaned_data["examination_year"], academic_year__end_date__year__gte=form.cleaned_data["examination_year"], number=int(row["term_number"]), is_deleted=False).first()
                    component = row["component"].strip().lower()
                    if not class_level or not term or component not in dict(UNEBContinuousAssessment.COMPONENTS):
                        raise ValueError(f"Row {row_number}: class, term or component is invalid.")
                    if request.user.role == "teacher" and not SubjectAllocation.objects.filter(teacher=request.user, subject=registration.subject, stream=candidate.student.stream, term=term, is_deleted=False, is_active=True).exists():
                        raise ValueError(f"Row {row_number}: subject/class is outside your teaching allocation.")
                    UNEBContinuousAssessment.objects.update_or_create(candidate_subject=registration, class_level=class_level, term=term, component=component, defaults={"raw_score": Decimal(row["raw_score"]), "maximum_score": Decimal(row["maximum_score"]), "evidence_reference": row["evidence_reference"].strip(), "notes": row.get("notes", "").strip(), "status": "draft", "entered_by": request.user})
                    count += 1
            messages.success(request, f"Imported {count} UNEB CA draft record(s). They must still be verified and approved.")
            return redirect("exams:uneb-ca-list")
        except Exception as exc:
            form.add_error("csv_file", str(exc))
            return render(request, self.template_name, {"form": form})


class UNEBCAImportTemplateView(AcademicStaffRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="uneb-ca-import-template.csv"'
        writer = csv.writer(response)
        writer.writerow(["index_number", "subject_code", "class_level", "term_number", "component", "raw_score", "maximum_score", "evidence_reference", "notes"])
        writer.writerow(["U9999/001", "ENG", "S4", "2", "activity_integration", "76", "100", "S4-ENG-T2-AOI-001", "Verified classroom evidence reference"])
        return response


class UNEBExportView(AcademicManagerRequiredMixin, View):
    @transaction.atomic
    def post(self, request):
        level = request.POST.get("level")
        year = request.POST.get("year")
        adapter = active_uneb_adapter(level, year) if level and year else None
        preflight = uneb_export_preflight(level, year, adapter)
        if not preflight.ready:
            for error in preflight.errors[:20]:
                messages.error(request, f"Export blocked: {error}")
            if len(preflight.errors) > 20:
                messages.error(request, f"{len(preflight.errors) - 20} additional issue(s) were omitted.")
            return redirect("exams:uneb-ca-list")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "UNEB CA pre-submission"
        headers = [column["header"] for column in uneb_export_columns(adapter)]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        selected = preflight.records
        for item in selected:
            sheet.append(uneb_export_row(item, adapter))
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 38)
        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()
        batch = UNEBExportBatch.objects.create(
            examination_level=level,
            examination_year=int(year),
            generated_by=request.user,
            record_count=len(selected),
            checksum=hashlib.sha256(payload).hexdigest(),
            adapter=adapter,
            format_version=adapter.version if adapter else "internal-pre-submission-v1",
            validation_report={"errors": preflight.errors, "warnings": preflight.warnings, "ready": preflight.ready},
        )
        batch.records.set(selected)
        UNEBContinuousAssessment.objects.filter(pk__in=[item.pk for item in selected]).update(status="exported")
        record_audit(request, "uneb_export", batch, f"Generated pre-submission export with {len(selected)} record(s)", {"checksum": batch.checksum})
        response = HttpResponse(payload, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{level.upper()}-{year}-CA-pre-submission.xlsx"'
        response["X-Export-Checksum"] = batch.checksum
        return response


class SetupTableView(AcademicManagerRequiredMixin, ModelTableView):
    pass


class SetupCreateView(AcademicManagerRequiredMixin, AuditedFormMixin, CreateView):
    template_name = "shared/form.html"
    audit_action = "create"

    def form_valid(self, form):
        if isinstance(form.instance, UNEBIntegrationAdapter) and form.instance.status == "active":
            form.instance.approved_by = self.request.user
        return super().form_valid(form)


class SetupUpdateView(AcademicManagerRequiredMixin, AuditedFormMixin, UpdateView):
    template_name = "shared/form.html"
    audit_action = "update"

    def form_valid(self, form):
        if isinstance(form.instance, UNEBIntegrationAdapter) and form.instance.status == "active":
            form.instance.approved_by = self.request.user
        return super().form_valid(form)
