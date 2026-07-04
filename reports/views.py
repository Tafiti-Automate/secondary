from io import BytesIO
from html import escape
from zipfile import ZipFile, ZIP_DEFLATED

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Avg, Count, Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views import View
from django.views.generic import DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak

from academics.models import SchoolProfile, Stream, Subject, SubjectAllocation, Term
from assessments.models import CompetencyAssessment
from attendance.models import AttendanceRecord
from config.crud import AuditedFormMixin
from config.audit import record_audit
from config.mixins import ACADEMIC_MANAGERS, ACADEMIC_STAFF, AcademicManagerRequiredMixin, AcademicStaffRequiredMixin
from students.models import Student
from .forms import ReportCardReviewForm, ReportGenerationForm, TranscriptGenerationForm
from .growth import learner_growth_profile
from .models import AcademicTranscript, ReportCard, ReportSubjectResult
from .services import generate_academic_transcript, generate_stream_reports


def build_report_pdf_bytes(card):
    """Render a self-contained report card PDF for individual and bulk downloads."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontSize=17, leading=21, textColor=colors.HexColor("#0f172a"), alignment=TA_CENTER, spaceAfter=3 * mm))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontSize=10, leading=13, textColor=colors.HexColor("#047857"), spaceBefore=4 * mm, spaceAfter=2 * mm))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=7.5, leading=10))
    styles.add(ParagraphStyle(name="Tiny", parent=styles["BodyText"], fontSize=6.8, leading=8.5))
    school = SchoolProfile.objects.filter(is_deleted=False).first()
    school_name = school.name if school else "Secondary School"
    story = [
        Paragraph(escape(school_name), styles["ReportTitle"]),
        Paragraph(escape(" · ".join(filter(None, [school.address if school else "", f"EMIS {school.emis_number}" if school and school.emis_number else "", f"UNEB {school.centre_number}" if school and school.centre_number else ""]))), styles["Small"]),
        Spacer(1, 2 * mm),
        Paragraph(f"<b>ACADEMIC PROGRESS REPORT</b> · {escape(str(card.term))}", styles["Heading2"]),
    ]
    student_data = [
        ["Learner", card.student.full_name, "Student no.", card.student.student_number or "—"],
        ["Admission no.", card.student.admission_number, "Class", str(card.stream)],
        ["Attendance", f"{card.attendance_percentage}%", "Position", f"{card.class_position or '—'} / {card.class_size or '—'}"],
    ]
    identity = Table(student_data, colWidths=[27 * mm, 62 * mm, 27 * mm, 50 * mm])
    identity.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecfdf5")), ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#ecfdf5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"), ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([identity, Paragraph("Subject performance", styles["Section"])])
    subject_rows = [["Subject", "Ongoing", "Summative", "Final", "Grade", "Descriptor / issue"]]
    for result in card.subject_results.filter(is_deleted=False).select_related("subject"):
        subject_rows.append([
            Paragraph(escape(result.subject.name), styles["Tiny"]),
            str(result.continuous_assessment_score) if result.ongoing_available else "Missing",
            str(result.examination_score) if result.summative_available else ("Missing" if result.summative_required else "Not due"),
            str(result.final_score) if result.final_score is not None else "Incomplete",
            result.grade or "—",
            Paragraph(escape(result.descriptor or result.missing_reason or result.teacher_remark or "—"), styles["Tiny"]),
        ])
    subjects = Table(subject_rows, repeatRows=1, colWidths=[37 * mm, 21 * mm, 23 * mm, 19 * mm, 14 * mm, 52 * mm])
    subjects.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]), ("PADDING", (0, 0), (-1, -1), 3.5),
    ]))
    story.append(subjects)

    def rating_text(queryset, label, value_getter):
        items = [f"<b>{escape(label(item))}</b>: {escape(value_getter(item))}" for item in queryset]
        return Paragraph("<br/>".join(items) if items else "No evidence recorded.", styles["Small"])

    story.append(Paragraph("Competencies, skills, values and learning outcomes", styles["Section"]))
    competency_text = rating_text(card.competency_ratings.filter(is_deleted=False).select_related("competency", "level"), lambda x: x.competency.name, lambda x: x.level.name)
    skill_text = rating_text(card.skill_ratings.filter(is_deleted=False).select_related("skill", "level"), lambda x: x.skill.name, lambda x: x.level.name)
    value_text = rating_text(card.value_ratings.filter(is_deleted=False).select_related("value", "level"), lambda x: x.value.name, lambda x: x.level.name)
    outcome_text = rating_text(card.learning_outcome_ratings.filter(is_deleted=False).select_related("outcome", "scale_level"), lambda x: x.outcome.statement, lambda x: x.display_level)
    evidence = Table([[competency_text, skill_text], [value_text, outcome_text]], colWidths=[83 * mm, 83 * mm])
    evidence.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("PADDING", (0, 0), (-1, -1), 5)]))
    story.extend([evidence, Paragraph("Portfolio", styles["Section"]), Paragraph(escape(card.portfolio_summary or "No verified portfolio evidence for this term."), styles["Small"])])
    comments = Table([
        [Paragraph("<b>Class teacher</b><br/>" + escape(card.teacher_remark or "No remark recorded."), styles["Small"]), Paragraph("<b>Director of Studies</b><br/>" + escape(card.dos_remark or "No remark recorded."), styles["Small"])],
        [Paragraph("<b>Headteacher</b><br/>" + escape(card.headteacher_comment or "No comment recorded."), styles["Small"]), Paragraph(f"<b>Promotion recommendation</b><br/>{escape(card.get_promotion_decision_display())}", styles["Small"])],
    ], colWidths=[83 * mm, 83 * mm])
    comments.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("PADDING", (0, 0), (-1, -1), 5)]))
    story.extend([Paragraph("Remarks and progression", styles["Section"]), comments])
    if school and school.report_footer:
        story.extend([Spacer(1, 3 * mm), Paragraph(escape(school.report_footer), styles["Tiny"])])
    doc.build(story)
    return buffer.getvalue()


def build_transcript_pdf_bytes(transcript):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=16 * mm, leftMargin=16 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    school = SchoolProfile.objects.filter(is_deleted=False).first()
    story = [
        Paragraph(escape(school.name if school else "Secondary School"), styles["Title"]),
        Paragraph("OFFICIAL ACADEMIC TRANSCRIPT", ParagraphStyle(name="TranscriptHeading", parent=styles["Heading2"], alignment=TA_CENTER, textColor=colors.HexColor("#047857"))),
        Spacer(1, 4 * mm),
    ]
    identity = Table([
        ["Learner", transcript.student.full_name, "Student no.", transcript.student.student_number or "—"],
        ["Admission no.", transcript.student.admission_number, "Serial", transcript.serial_number],
        ["Status", transcript.get_status_display(), "Issued", transcript.issued_at.strftime("%d %B %Y") if transcript.issued_at else "Draft"],
    ], colWidths=[28 * mm, 58 * mm, 28 * mm, 52 * mm])
    identity.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")), ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecfdf5")), ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#ecfdf5")), ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"), ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("PADDING", (0, 0), (-1, -1), 4)]))
    story.extend([identity, Spacer(1, 5 * mm)])
    rows = [["Academic year", "Term", "Class", "Subject", "Score", "Grade", "Points", "Remark"]]
    for entry in transcript.entries.filter(is_deleted=False).select_related("academic_year", "term", "class_level", "subject"):
        rows.append([str(entry.academic_year), entry.term.name if entry.term else "Year", str(entry.class_level), Paragraph(escape(entry.subject.name), styles["BodyText"]), str(entry.score), entry.grade or "—", str(entry.points or "—"), Paragraph(escape(entry.remark or "—"), styles["BodyText"])])
    table = Table(rows, repeatRows=1, colWidths=[23 * mm, 17 * mm, 15 * mm, 36 * mm, 16 * mm, 13 * mm, 13 * mm, 33 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("PADDING", (0, 0), (-1, -1), 3.5), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")])]))
    story.append(table)
    if transcript.purpose:
        story.extend([Spacer(1, 4 * mm), Paragraph(f"<b>Purpose:</b> {escape(transcript.purpose)}", styles["BodyText"])])
    story.extend([Spacer(1, 10 * mm), Paragraph("Authorised signature: ______________________________", styles["BodyText"]), Spacer(1, 3 * mm), Paragraph("This transcript is valid only when issued by the school and bearing the authorised signature/stamp.", styles["BodyText"])])
    doc.build(story)
    return buffer.getvalue()


def visible_report_cards(user):
    qs = ReportCard.objects.filter(is_deleted=False).select_related("student", "term", "stream").prefetch_related(
        "subject_results__subject", "competency_ratings__competency", "competency_ratings__level",
        "learning_outcome_ratings__outcome", "skill_ratings__skill", "skill_ratings__level",
        "value_ratings__value", "value_ratings__level",
    )
    if user.is_superuser or user.role in ACADEMIC_MANAGERS:
        return qs
    if user.role == "teacher":
        return qs.filter(Q(stream__class_teacher=user) | Q(stream__subject_allocations__teacher=user, stream__subject_allocations__is_deleted=False)).distinct()
    if user.role == "student":
        return qs.filter(student__user=user, status__in=("published", "locked"))
    if user.role == "parent":
        return qs.filter(Q(student__parent=user) | Q(student__guardian_links__guardian__user=user), status__in=("published", "locked")).distinct()
    return qs.none()


def visible_transcripts(user):
    qs = AcademicTranscript.objects.filter(is_deleted=False).select_related("student", "issued_by").prefetch_related("entries__subject", "entries__academic_year", "entries__term", "entries__class_level")
    if user.is_superuser or user.role in ACADEMIC_MANAGERS:
        return qs
    if user.role == "teacher":
        return qs.filter(Q(student__stream__class_teacher=user) | Q(student__stream__subject_allocations__teacher=user), status="issued").distinct()
    if user.role == "student":
        return qs.filter(student__user=user, status="issued")
    if user.role == "parent":
        return qs.filter(Q(student__parent=user) | Q(student__guardian_links__guardian__user=user), status="issued").distinct()
    return qs.none()


def visible_growth_students(user):
    qs = Student.objects.filter(is_deleted=False).select_related("stream", "stream__class_level")
    if user.is_superuser or user.role in ACADEMIC_MANAGERS:
        return qs
    if user.role == "teacher":
        return qs.filter(Q(stream__class_teacher=user) | Q(stream__subject_allocations__teacher=user)).distinct()
    if user.role == "student":
        return qs.filter(user=user)
    if user.role == "parent":
        return qs.filter(Q(parent=user) | Q(guardian_links__guardian__user=user)).distinct()
    return qs.none()


class LearnerGrowthProfileView(LoginRequiredMixin, DetailView):
    template_name = "reports/growth_profile.html"
    context_object_name = "student"

    def get_queryset(self):
        return visible_growth_students(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["growth"] = learner_growth_profile(self.object)
        return context


class ReportCardListView(LoginRequiredMixin, ListView):
    template_name = "reports/report_list.html"
    context_object_name = "reports"
    paginate_by = 24

    def get_queryset(self):
        qs = visible_report_cards(self.request.user)
        if self.request.GET.get("term"):
            qs = qs.filter(term_id=self.request.GET["term"])
        if self.request.GET.get("stream"):
            qs = qs.filter(stream_id=self.request.GET["stream"])
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({"terms": Term.objects.filter(is_deleted=False), "streams": Stream.objects.filter(is_deleted=False)})
        return context


class AcademicTranscriptListView(LoginRequiredMixin, ListView):
    template_name = "reports/transcript_list.html"
    context_object_name = "transcripts"
    paginate_by = 30

    def get_queryset(self):
        return visible_transcripts(self.request.user)


class AcademicTranscriptDetailView(LoginRequiredMixin, DetailView):
    template_name = "reports/transcript_detail.html"
    context_object_name = "transcript"

    def get_queryset(self):
        return visible_transcripts(self.request.user)


class GenerateTranscriptView(AcademicManagerRequiredMixin, View):
    template_name = "reports/transcript_generate.html"

    def get(self, request):
        return render(request, self.template_name, {"form": TranscriptGenerationForm()})

    def post(self, request):
        form = TranscriptGenerationForm(request.POST)
        if form.is_valid():
            try:
                transcript = generate_academic_transcript(form.cleaned_data["student"], request.user, form.cleaned_data["purpose"])
            except ValidationError as exc:
                form.add_error("student", "; ".join(exc.messages))
            else:
                messages.success(request, "Draft academic transcript generated.")
                return redirect("reports:transcript-detail", pk=transcript.pk)
        return render(request, self.template_name, {"form": form})


class IssueTranscriptView(AcademicManagerRequiredMixin, View):
    def post(self, request, pk):
        transcript = get_object_or_404(AcademicTranscript, pk=pk, is_deleted=False)
        if not transcript.entries.filter(is_deleted=False).exists():
            messages.error(request, "An empty transcript cannot be issued.")
        else:
            transcript.status = "issued"
            transcript.issued_by = request.user
            transcript.save()
            messages.success(request, "Transcript issued and made available to the learner and parent.")
        return redirect("reports:transcript-detail", pk=pk)


class AcademicTranscriptPDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        transcript = get_object_or_404(visible_transcripts(request.user), pk=pk)
        response = HttpResponse(build_transcript_pdf_bytes(transcript), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{transcript.serial_number}.pdf"'
        return response


class ReportCardDetailView(LoginRequiredMixin, DetailView):
    template_name = "reports/report_card.html"
    context_object_name = "report"

    def get_queryset(self):
        return visible_report_cards(self.request.user)


class GenerateReportsView(AcademicStaffRequiredMixin, View):
    template_name = "reports/generate.html"

    def get(self, request):
        form = ReportGenerationForm()
        self.scope_form(form, request.user)
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = ReportGenerationForm(request.POST)
        self.scope_form(form, request.user)
        if form.is_valid():
            cards = generate_stream_reports(form.cleaned_data["stream"], form.cleaned_data["term"], request.user, form.cleaned_data["assessment_policy"])
            if cards:
                record_audit(request, "generate_reports", cards[0], f"Generated {len(cards)} report card(s) for {form.cleaned_data['stream']}", {"count": len(cards)})
            messages.success(request, f"Generated {len(cards)} report card(s) and calculated class positions.")
            return redirect(f"{reverse('reports:list')}?term={form.cleaned_data['term'].pk}&stream={form.cleaned_data['stream'].pk}")
        return render(request, self.template_name, {"form": form})

    @staticmethod
    def scope_form(form, user):
        if user.role == "teacher" and not user.is_superuser:
            form.fields["stream"].queryset = form.fields["stream"].queryset.filter(Q(class_teacher=user) | Q(subject_allocations__teacher=user, subject_allocations__is_deleted=False)).distinct()
            form.fields["term"].queryset = form.fields["term"].queryset.filter(subject_allocations__teacher=user, subject_allocations__is_deleted=False).distinct()


class ReportReviewView(AcademicManagerRequiredMixin, AuditedFormMixin, UpdateView):
    model = ReportCard
    form_class = ReportCardReviewForm
    template_name = "shared/form.html"
    audit_action = "review"
    extra_context = {"page_title": "Review report card", "eyebrow": "Academic reporting", "submit_label": "Save review"}

    def get_success_url(self):
        return reverse("reports:detail", args=[self.object.pk])

    def form_valid(self, form):
        if form.instance.status == "draft":
            form.instance.status = "reviewed"
        return super().form_valid(form)


class ReportPublishView(AcademicManagerRequiredMixin, View):
    def post(self, request, pk):
        card = get_object_or_404(ReportCard, pk=pk, is_deleted=False)
        try:
            card.publish()
            record_audit(request, "publish", card, "Published report card to parent and student portals")
            messages.success(request, "Report card published to the parent and student portals.")
        except ValidationError as exc:
            messages.error(request, str(exc))
        return redirect("reports:detail", pk=pk)


class ReportBatchWorkflowView(AcademicManagerRequiredMixin, View):
    @transaction.atomic
    def post(self, request, operation):
        reports = ReportCard.objects.filter(pk__in=request.POST.getlist("report_ids"), is_deleted=False)
        completed = skipped = 0
        for card in reports:
            if operation == "review" and card.status == "draft":
                card.status = "reviewed"
                card.save(update_fields=["status", "updated_at"])
                completed += 1
            elif operation == "publish" and card.status in {"draft", "reviewed"} and not card.has_missing_marks:
                card.publish()
                completed += 1
            elif operation == "lock" and card.status == "published":
                card.status = "locked"
                card.save(update_fields=["status", "updated_at"])
                completed += 1
            else:
                skipped += 1
        if completed:
            first = reports.first()
            record_audit(request, f"batch_{operation}", first, f"Batch {operation}: {completed} report card(s)", {"completed": completed, "skipped": skipped})
        messages.success(request, f"{completed} report(s) processed; {skipped} skipped because of status or missing marks.")
        return redirect(request.META.get("HTTP_REFERER", reverse("reports:list")))


class MissingMarksView(AcademicStaffRequiredMixin, ListView):
    template_name = "reports/missing_marks.html"
    context_object_name = "results"
    paginate_by = 40

    def get_queryset(self):
        cards = visible_report_cards(self.request.user)
        qs = ReportSubjectResult.objects.filter(report_card__in=cards, is_deleted=False).filter(Q(ongoing_available=False) | Q(summative_required=True, summative_available=False)).select_related("report_card__student", "report_card__term", "report_card__stream", "subject")
        if self.request.GET.get("term"):
            qs = qs.filter(report_card__term_id=self.request.GET["term"])
        if self.request.GET.get("stream"):
            qs = qs.filter(report_card__stream_id=self.request.GET["stream"])
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({"terms": Term.objects.filter(is_deleted=False), "streams": Stream.objects.filter(is_deleted=False)})
        return context


class ReportCardPDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        card = get_object_or_404(visible_report_cards(request.user), pk=pk)
        response = HttpResponse(build_report_pdf_bytes(card), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="report-{card.student.admission_number.replace("/", "-")}.pdf"'
        return response


class BulkReportPDFZipView(AcademicStaffRequiredMixin, View):
    def get(self, request):
        term = get_object_or_404(Term, pk=request.GET.get("term"))
        stream = get_object_or_404(Stream, pk=request.GET.get("stream"))
        cards = visible_report_cards(request.user).filter(term=term, stream=stream).select_related("student", "term", "stream").prefetch_related(
            "subject_results__subject", "competency_ratings__competency", "competency_ratings__level",
            "skill_ratings__skill", "skill_ratings__level", "value_ratings__value",
            "value_ratings__level", "learning_outcome_ratings__outcome",
        )
        buffer = BytesIO()
        with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
            for card in cards.order_by("student__first_name", "student__last_name"):
                filename = f"{card.student.admission_number.replace('/', '-')}-{card.student.full_name.replace(' ', '_')}.pdf"
                archive.writestr(filename, build_report_pdf_bytes(card))
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{stream}-{term}-report-cards.zip"'
        return response


class PerformanceExcelView(AcademicStaffRequiredMixin, View):
    def get(self, request):
        term = get_object_or_404(Term, pk=request.GET.get("term"))
        stream = get_object_or_404(Stream, pk=request.GET.get("stream"))
        cards = ReportCard.objects.filter(term=term, stream=stream, is_deleted=False).prefetch_related("subject_results__subject")
        subjects = list(Subject.objects.filter(report_results__report_card__in=cards).distinct().order_by("name"))
        wb, ws = Workbook(), None
        ws = wb.active
        ws.title = "Class performance"
        ws.append(["Admission no.", "Student", *[s.name for s in subjects], "Average", "Position"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for card in cards.order_by("class_position"):
            score_map = {x.subject_id: float(x.final_score) if x.final_score is not None else "MISSING" for x in card.subject_results.filter(is_deleted=False)}
            ws.append([card.student.admission_number, card.student.full_name, *[score_map.get(s.pk, "") for s in subjects], float(card.average_score), card.class_position])
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{stream}-{term}-performance.xlsx"'
        return response


class AnalyticsDashboardView(AcademicStaffRequiredMixin, View):
    def get(self, request):
        term = Term.objects.filter(pk=request.GET.get("term")).first() or Term.objects.filter(is_current=True, is_deleted=False).first()
        subject_performance = ReportSubjectResult.objects.filter(report_card__term=term, is_deleted=False).values("subject__name").annotate(average=Avg("final_score"), students=Count("report_card__student", distinct=True)).order_by("-average") if term else []
        class_performance = ReportCard.objects.filter(term=term, is_deleted=False).values("stream__class_level__name", "stream__name").annotate(average=Avg("average_score"), students=Count("student")).order_by("stream__class_level__level") if term else []
        competency_performance = CompetencyAssessment.objects.filter(assessment__term=term, is_deleted=False, scale_level__isnull=False).values("competency__name").annotate(average=Avg("scale_level__numeric_value"), ratings=Count("id")).order_by("competency__display_order") if term else []
        attendance = AttendanceRecord.objects.filter(session__term=term, is_deleted=False).values("status").annotate(total=Count("id")) if term else []
        promotion = ReportCard.objects.filter(term=term, is_deleted=False).values("promotion_decision").annotate(total=Count("id")) if term else []
        teacher_performance = SubjectAllocation.objects.filter(term=term, is_deleted=False).values("teacher__first_name", "teacher__last_name").annotate(subjects=Count("subject", distinct=True), assessments=Count("subject__assessments", filter=Q(subject__assessments__term=term), distinct=True), results=Count("subject__examinations__results", filter=Q(subject__examinations__term=term), distinct=True)).order_by("teacher__first_name") if term else []
        return render(request, "reports/analytics.html", {"term": term, "terms": Term.objects.filter(is_deleted=False), "subject_performance": list(subject_performance), "class_performance": list(class_performance), "competency_performance": list(competency_performance), "attendance": list(attendance), "promotion": list(promotion), "teacher_performance": list(teacher_performance)})
